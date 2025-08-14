import streamlit as st
from PIL import Image, ImageOps, ImageFilter
import pytesseract
import pandas as pd
import re
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font

# 📍 Set Tesseract path (Streamlit Cloud default)
pytesseract.pytesseract.tesseract_cmd = "/usr/bin/tesseract"

# 📦 Image Preprocessing tuned for Mangal font
def preprocess_image(image):
    image = image.convert("L")
    image = ImageOps.autocontrast(image)
    image = image.filter(ImageFilter.SHARPEN)
    image = image.point(lambda x: 0 if x < 128 else 255, '1')
    return image

# 🧠 Deduplicate Column Names
def deduplicate_columns(columns):
    seen = {}
    new_cols = []
    for col in columns:
        col = str(col).strip()
        if col not in seen:
            seen[col] = 1
            new_cols.append(col)
        else:
            seen[col] += 1
            new_cols.append(f"{col}_{seen[col]}")
    return new_cols

# 🧹 Clean Hindi text
def clean_hindi_text(text):
    text = text.replace("।", ".")
    text = re.sub(r"[^\u0900-\u097F\s]", "", text)
    return text.strip()

# 🧾 Parse Table from OCR Data
def parse_table_from_data(ocr_df):
    ocr_df = ocr_df[ocr_df.text.notnull() & (ocr_df.text.str.strip() != "")]
    ocr_df = ocr_df.reset_index(drop=True)

    lines = ocr_df.groupby("line_num")
    rows = []
    for _, line in lines:
        words = line.sort_values("left")["text"].tolist()
        cleaned = [clean_hindi_text(w) for w in words]
        rows.append(cleaned)

    if not rows or len(rows) < 2:
        return pd.DataFrame()

    max_cols = max(len(row) for row in rows)
    rows = [row + [""] * (max_cols - len(row)) for row in rows]

    headers = deduplicate_columns(rows[0])
    df = pd.DataFrame(rows[1:], columns=headers)
    return df

# 📤 Convert DataFrame to CSV
def convert_df_to_csv(df):
    df.columns = deduplicate_columns([str(col) for col in df.columns])
    return df.to_csv(index=False).encode("utf-8")

# 📤 Export to styled Excel with Mangal font
def export_to_excel_with_mangal(df):
    wb = Workbook()
    ws = wb.active

    # Write headers
    for col_num, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num, value=col_name)
        cell.font = Font(name="Mangal", size=12, bold=True)

    # Write data
    for row_num, row in enumerate(df.itertuples(index=False), start=2):
        for col_num, value in enumerate(row, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=str(value))
            cell.font = Font(name="Mangal", size=12)

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# 🌟 Streamlit UI
st.set_page_config(page_title="Jamabandi OCR to Excel", page_icon="📄")
st.title("📄 Jamabandi OCR Table to Excel Converter")

uploaded_file = st.file_uploader("Upload a Jamabandi-style Hindi table image", type=["jpg", "jpeg", "png"])

if uploaded_file:
    image = Image.open(uploaded_file)
    st.image(image, caption="Uploaded Image", use_column_width=True)

    processed = preprocess_image(image)
    ocr_df = pytesseract.image_to_data(processed, lang="hin", output_type=pytesseract.Output.DATAFRAME)

    df = parse_table_from_data(ocr_df)

    if df.empty:
        st.warning("⚠️ Could not detect a table structure. Try a clearer image.")
        st.text_area("🔍 Raw OCR Output", "\n".join(ocr_df["text"].dropna().tolist()), height=200)
    else:
        st.subheader("✅ Extracted Table")
        st.dataframe(df)

        csv_data = convert_df_to_csv(df)
        st.download_button("📥 Download as CSV", data=csv_data, file_name="jamabandi_table.csv", mime="text/csv")

        excel_buffer = export_to_excel_with_mangal(df)
        st.download_button("📥 Download Styled Excel", data=excel_buffer, file_name="jamabandi_output.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
